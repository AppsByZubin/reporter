from __future__ import annotations

import re
import sys
from copy import copy
from decimal import Decimal
from pathlib import Path
from typing import Any

from common.constants import REPORT_COLUMNS, SECTION_BY_BOT
from utils.record_utils import clean_cell_value, decimal_or_none


AMOUNT_NEGATIVE_FONT_COLOR = "FFFF0000"
AMOUNT_POSITIVE_FONT_COLOR = "FF008000"
END_OF_DAY_TOTAL_LABEL = "End of the day total:"
END_OF_DAY_TOTAL_VALUE_OFFSET = 1
ORDER_ID_COLUMNS = {"broker_order_id", "exchange_order_id"}
REPORT_START_COL = 5
REPORT_START_COL_LETTER = "E"
REPORT_END_COL = REPORT_START_COL + len(REPORT_COLUMNS) - 1
REPORT_END_COL_LETTER = "P"
AMOUNT_COL = REPORT_START_COL + REPORT_COLUMNS.index("amount")
TOTAL_LABEL_COL = AMOUNT_COL - 1
ORDER_ID_MIN_COL_WIDTH = 24
ORDER_ID_ROW_HEIGHT = 30


def write_report(
    template_path: Path,
    output_path: Path,
    bots: list[str],
    report_data: dict[str, tuple[list[dict[str, Any]], str]],
) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
    except ImportError as exc:
        raise RuntimeError(
            "Missing dependency openpyxl. Install with: python -m pip install -r requirements.txt"
        ) from exc

    wb = load_workbook(template_path)
    ws = wb.active
    unmerge_template_ranges(ws)
    format_order_id_columns(ws)

    offset = 0
    end_of_day_total = Decimal("0")
    sections_to_merge: list[tuple[int, int, int, int]] = []

    for bot in template_ordered_bots(bots):
        key = bot.lower()
        if key not in SECTION_BY_BOT:
            print(f"Skipping {bot}: no section for this bot in the template.", file=sys.stderr)
            continue

        rows, _observation = report_data.get(bot, ([], ""))
        section = prepare_section(ws, SECTION_BY_BOT[key], max(1, len(rows)), offset)
        offset = int(section["offset"])

        title_row = int(section["title_row"])
        header_row = int(section["header_row"])
        data_start = int(section["data_start"])
        data_end = int(section["data_end"])
        total_row = int(section["total_row"])
        observation_start = int(section["observation_start"])
        observation_end = int(section["observation_end"])

        ws.cell(title_row, REPORT_START_COL).value = bot.capitalize()
        clear_cells(ws, data_start, data_end, REPORT_START_COL, REPORT_END_COL)

        for row_index, row in enumerate(rows, start=data_start):
            for col_index, column in enumerate(REPORT_COLUMNS, start=REPORT_START_COL):
                cell = ws.cell(row_index, col_index)
                cell.value = clean_report_cell_value(column, row.get(column))
                if column in ORDER_ID_COLUMNS:
                    wrap_cell(cell)
                    if isinstance(cell.value, str) and "\n" in cell.value:
                        ensure_row_height(ws, row_index, ORDER_ID_ROW_HEIGHT)
                if column == "amount":
                    color_amount_font(cell, row.get(column))

        ws.cell(total_row, TOTAL_LABEL_COL).value = "Total:"
        total_amount = sum_amounts(rows)
        end_of_day_total += total_amount
        total_cell = ws.cell(total_row, AMOUNT_COL)
        total_cell.value = clean_cell_value(total_amount)
        color_amount_font(total_cell, total_amount)

        clear_cells(ws, observation_start, observation_end, REPORT_START_COL, REPORT_END_COL)
        ws.cell(observation_start, REPORT_START_COL).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

        table_name = str(section["table"])
        if table_name in ws.tables:
            ws.tables[table_name].ref = (
                f"{REPORT_START_COL_LETTER}{header_row}:"
                f"{REPORT_END_COL_LETTER}{total_row}"
            )

        sections_to_merge.append(
            (title_row, title_row, REPORT_START_COL, REPORT_END_COL)
        )
        sections_to_merge.append(
            (observation_start, observation_end, REPORT_START_COL, REPORT_END_COL)
        )

    for start_row, end_row, start_col, end_col in sections_to_merge:
        ws.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=end_row,
            end_column=end_col,
        )

    write_end_of_day_total(ws, end_of_day_total)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def unmerge_template_ranges(ws: Any) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        range_text = str(merged_range)
        if re.fullmatch(r"E\d+:P\d+", range_text):
            ws.unmerge_cells(range_text)


def copy_row_style(
    ws: Any,
    source_row: int,
    target_row: int,
    start_col: int,
    end_col: int,
) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(start_col, end_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)


def prepare_section(
    ws: Any,
    section: dict[str, Any],
    rows_needed: int,
    offset: int,
) -> dict[str, int | str]:
    available = int(section["data_end"]) - int(section["data_start"]) + 1
    base_offset = offset
    title_row = int(section["title_row"]) + base_offset
    header_row = int(section["header_row"]) + base_offset
    data_start = int(section["data_start"]) + base_offset
    total_row = int(section["total_row"]) + base_offset
    observation_start = int(section["observation_start"]) + base_offset
    observation_end = int(section["observation_end"]) + base_offset
    extra_rows = max(0, rows_needed - available)

    if extra_rows:
        ws.insert_rows(total_row, extra_rows)
        for row in range(total_row, total_row + extra_rows):
            copy_row_style(ws, data_start, row, REPORT_START_COL, REPORT_END_COL)
        offset += extra_rows
        total_row += extra_rows
        observation_start += extra_rows
        observation_end += extra_rows

    return {
        "offset": offset,
        "title_row": title_row,
        "header_row": header_row,
        "data_start": data_start,
        "data_end": total_row - 1,
        "total_row": total_row,
        "observation_start": observation_start,
        "observation_end": observation_end,
        "table": str(section["table"]),
    }


def template_ordered_bots(bots: list[str]) -> list[str]:
    return sorted(
        bots,
        key=lambda bot: SECTION_BY_BOT.get(
            bot.lower(),
            {"title_row": 999_999},
        )["title_row"],
    )


def clear_cells(ws: Any, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).value = None


def color_amount_font(cell: Any, value: Any) -> None:
    amount = decimal_or_none(value)
    if amount is None:
        return

    if amount < 0:
        color = AMOUNT_NEGATIVE_FONT_COLOR
    elif amount > 0:
        color = AMOUNT_POSITIVE_FONT_COLOR
    else:
        return

    font = copy(cell.font)
    font.color = color
    cell.font = font


def clean_report_cell_value(column: str, value: Any) -> Any:
    if column in ORDER_ID_COLUMNS:
        return clean_text_cell_value(value)
    return clean_cell_value(value)


def clean_text_cell_value(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def wrap_cell(cell: Any) -> None:
    alignment = copy(cell.alignment)
    alignment.wrap_text = True
    alignment.vertical = "top"
    cell.alignment = alignment


def ensure_row_height(ws: Any, row_index: int, height: int) -> None:
    current = ws.row_dimensions[row_index].height or 0
    if current < height:
        ws.row_dimensions[row_index].height = height


def format_order_id_columns(ws: Any) -> None:
    from openpyxl.utils import get_column_letter

    for column in ORDER_ID_COLUMNS:
        col_index = REPORT_START_COL + REPORT_COLUMNS.index(column)
        col_letter = get_column_letter(col_index)
        current = ws.column_dimensions[col_letter].width or 0
        if current < ORDER_ID_MIN_COL_WIDTH:
            ws.column_dimensions[col_letter].width = ORDER_ID_MIN_COL_WIDTH


def sum_amounts(rows: list[dict[str, Any]]) -> Decimal:
    return sum(
        (
            amount
            for amount in (decimal_or_none(row.get("amount")) for row in rows)
            if amount is not None
        ),
        Decimal("0"),
    )


def write_end_of_day_total(ws: Any, total_amount: Decimal) -> None:
    row = find_end_of_day_total_row(ws)
    if row is None:
        return

    label_cell = ws.cell(row, REPORT_START_COL)
    label_cell.value = END_OF_DAY_TOTAL_LABEL

    total_cell = ws.cell(row, label_cell.column + END_OF_DAY_TOTAL_VALUE_OFFSET)
    total_cell.value = clean_cell_value(total_amount)
    color_amount_font(total_cell, total_amount)


def find_end_of_day_total_row(ws: Any) -> int | None:
    target = END_OF_DAY_TOTAL_LABEL.lower()
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.strip().lower() == target:
                return int(cell.row)
    return None
