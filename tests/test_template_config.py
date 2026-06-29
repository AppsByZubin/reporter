from unittest import TestCase

from common.constants import DEFAULT_BOT_LIST, DEFAULT_TEMPLATE, SECTION_BY_BOT
from utils.config_utils import read_bot_list



class TemplateConfigTests(TestCase):
    def test_configured_bots_have_template_sections(self) -> None:
        configured_bots = {bot.lower() for bot in read_bot_list(DEFAULT_BOT_LIST)}

        self.assertFalse(configured_bots - set(SECTION_BY_BOT))

    def test_sections_match_template(self) -> None:
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        workbook = load_workbook(DEFAULT_TEMPLATE)
        worksheet = workbook.active

        for bot, section in SECTION_BY_BOT.items():
            with self.subTest(bot=bot):
                self.assertEqual(
                    worksheet.cell(section["title_row"], 5).value.strip().lower(),
                    bot,
                )
                self.assertEqual(
                    worksheet.cell(section["header_row"], 5).value,
                    "trade_id",
                )
                self.assertEqual(
                    worksheet.cell(section["total_row"], 15).value,
                    "Total:",
                )
                self.assertIn(section["table"], worksheet.tables)
